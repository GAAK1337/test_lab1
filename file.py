import pandas as pd
import wget
import numpy
from numpy import unique
wget.download("https://drive.google.com/uc?export=download&id=1Bo5Oili5dAvWDSzAZXjzgjS71IrmLWun")
data =pd.read_excel("lab_pi_101.xlsx")
df=pd.DataFrame(data)

#num_rows = df.shape[0]
#print('Количество оценок :',num_rows)
a=len(df)
#print('Количество оценок :',a)
from openpyxl import load_workbook

wb = load_workbook('lab_pi_101.xlsx')

sheet = wb.get_sheet_by_name('Лист1')
kol=0
for i in range(1, a):
    y=(sheet.cell(row=i, column=11).value)
    if y=='ПИ101':
        kol=kol+1
print('В исходном датасете содержалось:',a ,'оценок, из них:',kol, 'оценок относятся к группе ПИ101')
     
nomer=pd.unique(df[['Личный номер студента']]. values.ravel ())
nom=list()
for i in range(1, a):
    
    n=(sheet.cell(row=i, column=10).value)
    g=(sheet.cell(row=i, column=11).value)
    if g=='ПИ101':
       nom.append(n)

print('В датасете находятся оценки',len (nomer), 'студентов со следующими личными номерами из группы ПИ101:', (unique(nom)))

uniques=pd.unique(df[['Уровень контроля']]. values.ravel ())
print('Используемые формы контроля:',uniques)

god=pd.unique(df[['Год']]. values.ravel ())
print('Данные представлены по следующим учебным годам:',god)
